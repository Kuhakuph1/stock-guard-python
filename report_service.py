from openpyxl import Workbook

from openpyxl.styles import (
    PatternFill,
    Font
)
from openpyxl.utils import (
    get_column_letter
)

def format_worksheet(ws):

    header_fill = PatternFill(
        "solid",
        fgColor="00B0F0"
    )

    header_font = Font(
        bold=True
    )

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font

    zebra_fill = PatternFill(
        "solid",
        fgColor="EAEAEA"
    )

    for row in range(
        2,
        ws.max_row + 1
    ):

        if row % 2 == 0:

            for cell in ws[row]:

                cell.fill = zebra_fill

    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:

                if len(
                    str(cell.value)
                ) > max_length:

                    max_length = len(
                        str(cell.value)
                    )

            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 5

def create_low_stock_excel(items):

    wb = Workbook()

    ws = wb.active

    ws.title = "LOW STOCK"

    ws.append([
        "Kode",
        "Nama",
        "Stock",
        "Min Stock"
    ])

    for item in items:

        ws.append([
            item["kode"],
            item["nama"],
            item["stock"],
            item["min"]
        ])

    filename = "low_stock_report.xlsx"

    format_worksheet(ws)

    wb.save(filename)

    return filename


def create_low_stock_txt(items):

    filename = "low_stock_report.txt"

    with open(
        filename,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(
            "LOW STOCK REPORT\n"
        )

        f.write(
            "=" * 50 + "\n\n"
        )

        for item in items:

            f.write(
                f"Kode : {item['kode']}\n"
            )

            f.write(
                f"Nama : {item['nama']}\n"
            )

            f.write(
                f"Stock : {item['stock']}\n"
            )

            f.write(
                f"Min Stock : {item['min']}\n"
            )

            f.write(
                "-" * 30 + "\n"
            )

    return filename
def create_stock_excel(
    lokasi,
    brand,
    jenis,
    stock_data
):

    wb = Workbook()

    ws = wb.active

    ws.title = "STOCK"

    ws.append([
        "Lokasi",
        "Merk",
        "Jenis",
        "Kode Barang",
        "Qty"
    ])

    for kode, qty in stock_data.items():

        ws.append([
            lokasi,
            brand,
            jenis,
            kode,
            qty
        ])

    filename = (
        f"stock_{brand}_{jenis}.xlsx"
    )
    format_worksheet(ws)
    wb.save(
        filename
    )

    return filename


def create_stock_txt(
    lokasi,
    brand,
    jenis,
    stock_data
):

    filename = (
        f"stock_{brand}_{jenis}.txt"
    )

    with open(
        filename,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(
            "STOCK REPORT\n"
        )

        f.write(
            "=" * 50 + "\n\n"
        )

        f.write(
            f"Lokasi : {lokasi}\n"
        )

        f.write(
            f"Merk : {brand}\n"
        )

        f.write(
            f"Jenis : {jenis}\n\n"
        )

        for kode, qty in stock_data.items():

            f.write(
                f"{kode} : {qty}\n"
            )

    return filename

def create_master_stock_excel(
    items
):

    wb = Workbook()

    ws = wb.active

    ws.title = "MASTER STOCK"

    ws.append([
        "Merk",
        "Kode",
        "Nama",
        "Class",
        "Min",
        "Max",
        "Actual",
        "Status"
    ])

    for item in items:

        ws.append([
            item["merk"],
            item["kode"],
            item["nama"],
            item["class"],
            item["min_stock"],
            item["max_stock"],
            item["actual_stock"],
            item["status"]
        ])

    filename = (
        "master_stock_report.xlsx"
    )
    format_worksheet(ws)
    wb.save(filename)

    return filename


def create_master_stock_txt(
    items
):

    filename = (
        "master_stock_report.txt"
    )

    with open(
        filename,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(
            "MASTER STOCK REPORT\n\n"
        )

        for item in items:

            f.write(
                f"{item['kode']} | "
                f"{item['nama']} | "
                f"{item['actual_stock']}\n"
            )

    return filename